from openpyxl import load_workbook
wb = load_workbook('test1.xlsx');
print(wb.sheetnames);
ws = wb.active;
# ws.auto_filter.ref = "A1:AE50";
# ws.auto_filter.add_filter_column(3,["7"]);
# wb.save("filtered1.xlsx")
count0 = 0;
count1 = 0;
count2 = 0;
strRDM0 = "";
strRDM1 = "";
strRDM2 = "";
for row in ws.iter_rows():
    # print(type(row[3]));
    test_value = row[3].value;
    if (test_value == "0")|(test_value =="5")|(test_value =="7"):
        # print('0');
        state = row[6].value;
        task = row[7].value;
        zr = row[10].value;
        sh = row[11].value;
        if (test_value == "0"):
            count0 += 1;
            strline = str(count0) + '.' + task + ';责任人: ' + zr + ',审核人: ' + sh + ';状态: ' + state;
            strRDM0 += (strline + '\n');
        elif (test_value == "5"):
            count1 += 1;
            strline = str(count1) + '.' + task + ';责任人: ' + zr + ',审核人: ' + sh + ';状态: ' + state;
            strRDM1 += (strline + '\n');
        elif (test_value == "7"):
            count2 += 1;
            strline = str(count2) + '.' + task + ';责任人: ' + zr + ',审核人: ' + sh + ';状态: ' + state;
            strRDM2 += (strline + '\n');

        # print(strline);
# print(strRDM0);
# print(strRDM1);
# print(strRDM2);
strout = "一、今日需关闭RDM任务，"+str(count0)+"个：\n"+strRDM0+"二、倒计时1天需关闭RDM任务，"+str(count1)+"个：\n"+strRDM1+\
"三、倒计时2天需关闭RDM任务，"+str(count2)+"个：\n"+strRDM2+'\n'
# print(strout);

data = open("out.txt", 'w+')
print(strout, file=data)
data.close()
print('1');
